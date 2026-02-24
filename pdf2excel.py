import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
from PIL import Image, ImageDraw, ImageFont
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
import math

# ── Configuración ────────────────────────────────────────────────────────────
COLS, ROWS = 4, 4
VOUCHERS_PER_PAGE = COLS * ROWS  # 16 por página

VOUCHER_W, VOUCHER_H = 500, 500  # Tamaño fijo por voucher
MARGIN = 15

PAGE_W = MARGIN + COLS * (VOUCHER_W + MARGIN)
PAGE_H = MARGIN + ROWS * (VOUCHER_H + MARGIN)

# Posición relativa del código dentro del voucher
CODE_REL_X = 0.50
CODE_REL_Y = 0.560

# ── Funciones ────────────────────────────────────────────────────────────────

def seleccionar_pdf_codigos():
    ruta = filedialog.askopenfilename(
        title="Selecciona el PDF con los códigos",
        filetypes=[("PDF files", "*.pdf")]
    )
    if ruta:
        entrada_codigos.set(ruta)

def seleccionar_pdf_voucher():
    ruta = filedialog.askopenfilename(
        title="Selecciona el PDF del diseño del voucher",
        filetypes=[("PDF files", "*.pdf")]
    )
    if ruta:
        entrada_voucher.set(ruta)

def extraer_codigos(ruta_pdf):
    with pdfplumber.open(ruta_pdf) as pdf:
        paginas = []
        for page in pdf.pages:
            texto = page.extract_text()
            if texto:
                paginas.append(texto.strip())
        texto_completo = " ".join(paginas)
        texto_completo = re.sub(r'\s+', ' ', texto_completo)
    return re.findall(r'\b\d{5}-\d{5}\b', texto_completo)

def generar_excel():
    ruta_codigos = entrada_codigos.get()
    if not ruta_codigos:
        messagebox.showwarning("Advertencia", "Selecciona el PDF con los códigos.")
        return

    codigos = extraer_codigos(ruta_codigos)
    if not codigos:
        messagebox.showinfo("Info", "No se encontraron códigos en el PDF.")
        return

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws['A1'] = "#"
    ws['B1'] = "Código"
    for cell in [ws['A1'], ws['B1']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, codigo in enumerate(codigos, 1):
        ws.append([i, codigo])
        ws.cell(row=i+1, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=i+1, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=i+1, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=i+1, column=2).font = Font(name="Arial", size=10)
        if i % 2 == 0:
            fill = PatternFill("solid", start_color="D6E4F0")
            ws.cell(row=i+1, column=1).fill = fill
            ws.cell(row=i+1, column=2).fill = fill

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18

    descargas = os.path.join(os.path.expanduser("~"), "Downloads")
    ruta_salida = os.path.join(descargas, "vouchersWifi.xlsx")
    wb.save(ruta_salida)

    messagebox.showinfo("¡Listo!", f"Se exportaron {len(codigos)} códigos.\nGuardado en:\n{ruta_salida}")

def generar_pdf():
    ruta_codigos = entrada_codigos.get()
    ruta_voucher = entrada_voucher.get()

    if not ruta_codigos or not ruta_voucher:
        messagebox.showwarning("Advertencia", "Selecciona ambos archivos PDF.")
        return

    codigos = extraer_codigos(ruta_codigos)
    if not codigos:
        messagebox.showinfo("Info", "No se encontraron códigos en el PDF.")
        return

    # Convertir voucher PDF a imagen
    try:
        imagenes = convert_from_path(ruta_voucher, dpi=150, poppler_path=r"C:\poppler-25.12.0\Library\bin")
        template = imagenes[0].convert("RGBA")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo convertir el voucher:\n{e}")
        return

    template_resized = template.resize((VOUCHER_W, VOUCHER_H), Image.LANCZOS)

    # Configurar fuente
    font_size = max(10, int(VOUCHER_H * 0.065))
    try:
        font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
    except:
        font = ImageFont.load_default()

    # Generar páginas
    total_paginas = math.ceil(len(codigos) / VOUCHERS_PER_PAGE)
    paginas_img = []

    for num_pagina in range(total_paginas):
        codigos_pagina = codigos[num_pagina * VOUCHERS_PER_PAGE:(num_pagina + 1) * VOUCHERS_PER_PAGE]
        page_img = Image.new("RGB", (PAGE_W, PAGE_H), "white")

        for idx, codigo in enumerate(codigos_pagina):
            col = idx % COLS
            row = idx // COLS

            x = MARGIN + col * (VOUCHER_W + MARGIN)
            y = MARGIN + row * (VOUCHER_H + MARGIN)

            voucher = template_resized.copy().convert("RGBA")
            draw = ImageDraw.Draw(voucher)

            # Cubrir código original con blanco
            code_x = int(VOUCHER_W * CODE_REL_X)
            code_y = int(VOUCHER_H * CODE_REL_Y)
            rect_h = int(VOUCHER_H * 0.09)
            rect_w = int(VOUCHER_W * 0.75)
            draw.rectangle([
                code_x - rect_w // 2, code_y - rect_h // 2,
                code_x + rect_w // 2, code_y + rect_h // 2
            ], fill="white")

            # Escribir nuevo código centrado
            bbox = draw.textbbox((0, 0), codigo, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
            draw.text(
                (code_x - text_w // 2, code_y - text_h // 2),
                codigo, fill="black", font=font
            )

            page_img.paste(voucher.convert("RGB"), (x, y))

        # Líneas de corte
        draw_page = ImageDraw.Draw(page_img)
        for col in range(1, COLS):
            lx = MARGIN + col * (VOUCHER_W + MARGIN) - MARGIN // 2
            draw_page.line([(lx, 0), (lx, PAGE_H)], fill="#bbbbbb", width=1)
        for row in range(1, ROWS):
            ly = MARGIN + row * (VOUCHER_H + MARGIN) - MARGIN // 2
            draw_page.line([(0, ly), (PAGE_W, ly)], fill="#bbbbbb", width=1)

        paginas_img.append(page_img)

    descargas = os.path.join(os.path.expanduser("~"), "Downloads")
    ruta_salida = os.path.join(descargas, "vouchersWifi.pdf")

    paginas_img[0].save(
        ruta_salida,
        save_all=True,
        append_images=paginas_img[1:],
        resolution=150
    )

    messagebox.showinfo(
        "¡Listo!",
        f"Se generaron {len(codigos)} vouchers en {total_paginas} páginas.\nGuardado en:\n{ruta_salida}"
    )

# ── Interfaz ─────────────────────────────────────────────────────────────────
app = tk.Tk()
app.title("Generador de Vouchers WiFi")
app.geometry("500x230")
app.resizable(False, False)

entrada_codigos = tk.StringVar()
entrada_voucher = tk.StringVar()

tk.Label(app, text="PDF con códigos:").grid(row=0, column=0, padx=10, pady=15, sticky="w")
tk.Entry(app, textvariable=entrada_codigos, width=38).grid(row=0, column=1, padx=5)
tk.Button(app, text="Buscar", command=seleccionar_pdf_codigos).grid(row=0, column=2, padx=5)

tk.Label(app, text="PDF diseño voucher:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
tk.Entry(app, textvariable=entrada_voucher, width=38).grid(row=1, column=1, padx=5)
tk.Button(app, text="Buscar", command=seleccionar_pdf_voucher).grid(row=1, column=2, padx=5)

tk.Label(app, text="(Solo necesario para generar PDF)", fg="gray", font=("Arial", 8)).grid(
    row=2, column=1, sticky="w", padx=5
)

# Botones de acción
frame_botones = tk.Frame(app)
frame_botones.grid(row=3, column=0, columnspan=3, pady=20)

tk.Button(
    frame_botones, text="📄  Generar PDF con diseño",
    command=generar_pdf,
    bg="#1F4E79", fg="white", width=22, height=2
).pack(side="left", padx=10)

tk.Button(
    frame_botones, text="📊  Generar Excel de códigos",
    command=generar_excel,
    bg="#1E6B2E", fg="white", width=22, height=2
).pack(side="left", padx=10)

app.mainloop()